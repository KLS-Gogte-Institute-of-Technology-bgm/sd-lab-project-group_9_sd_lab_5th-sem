#use your terminals.. make sure u have python and pip installed... run these two commands.. then run the actual code---- command to run it using ur terminal --- python3 barcode-reader.py
#pip install opencv-python
#pip install pyzbar
#pip install xlsxwriter
import cv2 
from pyzbar import pyzbar
import xlsxwriter 
import openpyxl

def read_barcodes(frame):
    barcodes = pyzbar.decode(frame)
    
    
    for barcode in barcodes:
        x, y, w, h = barcode.rect #barcode detected
        
        barcode_info = barcode.data.decode('utf-8') #info of barcode
        #print(barcode_info) 
        cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 255, 0), 1) 
        font = cv2.FONT_HERSHEY_DUPLEX
        cv2.putText(frame, barcode_info, (x + 6, y - 6), font, 2.0, (255, 255, 255), 1)
        
        content.append(barcode_info)

    return frame

def cam():
    
    print("[INFO] starting camera...")
    cam = cv2.VideoCapture(0) #turn on the main cam
    ret, frame = cam.read()
    
    while ret:
        ret, frame = cam.read()
        frame = read_barcodes(frame) 
        cv2.imshow('Online Attendence Scanner', frame)
        
        if cv2.waitKey(1) & 0xFF == 27:
            break
    #stores unique uids
    for uid in content:
        if uid not in uidlist:
            uidlist.append(uid)
    print(uidlist) # stores the unique uids generated from the scan
    
    workbook = xlsxwriter.Workbook('AL.xlsx') #can name the file as a timestamp
    worksheet = workbook.add_worksheet() 
    worksheet.write('A1', 'USN') 
    worksheet.write('B1', '24/12/20') #date is hard coded.... needs to be fetched from faculty login
    r, c= 1, 0
    #write into excel sheet.... for loop is for uidlist... change it to usn by mapping
    for item in uidlist:
        worksheet.write(r, c, item)
        worksheet.write(r, c+1, 'P')
        r+=1
    workbook.close() 
    cam.release()
    cv2.destroyAllWindows()


# def lol():

    
#     wb = openpyxl.load_workbook('AL.xlsx')
#     ws = wb.active

#     wb1 = openpyxl.load_workbook('Students.xlsx')
#     ws1 = wb1.active

#     from openpyxl import load_workbook
#     wb=load_workbook(filename='Students.xlsx')
#     sheet_ranges=wb['Sheet1']
#     #print(sheet_ranges['A2'].value)
#     #print(sheet_ranges['A5'].value)




# # for (col, col_1) in zip(ws.iter_cols(), ws1.iter_cols()):
# #     for (cell, cell_1) in zip(col, col_1):
# #        if cell.value != cell_1.value:
# #            print(str(cell.value) + ' is not equal to ' + str(cell_1.value) + ' ' + str(cell.coordinate)) 
# # ws1 = wb1.get_sheet_by_name("Sheet1")
# # ws2 = wb1.get_sheet_by_name("Sheet2")
#     c,r=6,1
#     for num in ws['A']:
#         print(num.value)
#         for target in ws1['C']:
#             if num.value == target.value:
#                 cell_obj=ws1.cell(row=r, column=c)
#                 cell_obj.value='P'
#                 print('match found')
                
#                 wb1.save('Students.xlsx')
        
#         r=r+1

def lol():
    
    wb = openpyxl.load_workbook('AL.xlsx')
    ws = wb.active

    wb1 = openpyxl.load_workbook('Students.xlsx')
    ws1 = wb1.active

    l=ws1.max_row
    a=ws.max_row
    for num in range (1,a+1):
        al_obj= ws.cell(row = num, column = 1)
        p=al_obj.value
        for target in range (1,l+1):
            cmp_obj= ws1.cell(row= target, column= 2)
            t=cmp_obj.value
            if p == t:
                cell_obj=ws1.cell(row=target, column=5)
                cell_obj.value='P'
            
                #print(cell_obj)
                wb1.save('Students.xlsx')
                break
    # for i in range (1,l+1):
    #     obj= ws1.cell(row= i, column= 5)
    #     t=obj.value
    #     if t=='':
    #         o=ws1.cell(row=i, column=5)
    #         o.value='Ab'


if __name__=='__main__':
    content = []
    uidlist=[]
    cam()
    lol()
